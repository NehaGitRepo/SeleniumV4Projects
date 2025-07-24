import time

from selenium import webdriver

from selenium.webdriver.common.by import By
from openpyxl import load_workbook
#opening excel file
workbook = load_workbook('testdata.xlsx')
#selecting excel sheet
sheet = workbook.active
driver = webdriver.Firefox()
driver.maximize_window()

for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True):
    username = row[0]
    password = row[1]
    driver.get("https://www.saucedemo.com/")
    time.sleep(2)

    driver.find_element(By.ID,"user-name").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "login-button").click()
    time.sleep(2)
    driver.find_element(By.XPATH,"//button[@id='react-burger-menu-btn']").click()
    time.sleep(2)
    driver.find_element(By.XPATH,"//a[@id='logout_sidebar_link']").click()
    time.sleep(2)
driver.quit()